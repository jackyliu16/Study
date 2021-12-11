"""
Author： 刘逸珑
Time：   2021/12/8 21:09
References:
    https://www.bilibili.com/video/BV1X54y1R7cu?from=search&seid=17433072440613079399&spm_id_from=333.337.0.0
    https://blog.csdn.net/heyuexianzi/article/details/76851377
    https://github.com/amueller/word_cloud/blob/master/examples/masked.py
"""
import openpyxl
from wordcloud import WordCloud # 不知道为什么在3.9中会出现安装失败
import matplotlib.pyplot as plt
import numpy as np
import PIL.Image as Image
import cv2
# background_image = np.array(Image.open("中国地图.jpg"))

def generate_pic(frequency,filename,savename):
    if filename != None:
        # 这个地方的作用在于将灰度50以下的东西设置为 0，以契合下文笼罩图的要求
        img_gary = cv2.imread(filename, cv2.IMREAD_GRAYSCALE)
        ret, img = cv2.threshold(img_gary, 50, 255, cv2.THRESH_BINARY)  # 有点不是很能理解为什么没有办法直接将numpy的东西赋值给backgroundimg
        cv2.imwrite('temp.png', img)
        background_image = np.array(Image.open("temp.png"))
    else:
        background_image = np.zeros((1080,1920),np.int8)
    # TODO 这个地方存在转化问题，在讲图片转化为矩阵的时候出现异常的边框
    wordcloud = WordCloud(font_path="HGKT_CNKI.TTF",
                          background_color="white",
                          mask = background_image,
                          # If mask is not None, width and height will be ignored and the shape of mask will be used instead.
                          # All white (#FF or #FFFFFF) entries will be considerd "masked out" while other entries will be free to draw on.
                          contour_width = 1,
                          # max_words = 300,
                          max_font_size = 600,
                          # contour_color = 'blue',
                          width=1920, height=1080)
    # 生成词云
    wordcloud.generate_from_frequencies(frequency)
    # 保存词云
    wordcloud.to_file("{}.png".format(savename))

# Read Data
wb = openpyxl.load_workbook('data.xlsx')
# 获取工作表
ws = wb['国内疫情']
frequencyIn = {}
for row in ws.values:
    if row[0] == "省份":
        pass
    else:
        frequencyIn[row[0]] = float(row[1])      # 生成一个以省份为key，确诊人数为value的字典

generate_pic(frequencyIn, savename="国内疫情词云", filename=None)



frequencyOut = {}
# 加入中国确诊人数
ChinaAccount = 0
for key,value in frequencyIn.items():
    ChinaAccount += float(value)
frequencyOut['中国'] = ChinaAccount

# wb.sheetnames = 工作表名称
sheet_name = wb.sheetnames
for each_sheet in sheet_name:     # 由于不同洲的数据被保存在不同的工作簿中，因此需要先对于工作表进行遍历
    if "洲" in each_sheet:
        ws = wb[each_sheet]
        for row in ws.values:
            if row[0] == "国家":
                pass
            else:
                frequencyOut[row[0]] = float(row[1])

generate_pic(frequencyOut,filename='881259.png',savename="世界疫情词云")
